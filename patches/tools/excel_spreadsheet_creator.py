"""
title: Excel Spreadsheet Creator
author: AI Stack
description: >
  Creates formatted Excel (.xlsx) spreadsheets from structured data.
  Supports multiple sheets, styled headers, data formatting, and totals rows.
  Registers the file with Open WebUI and returns a proper download link.
version: 3.1.0
requirements: openpyxl
"""

import os
import re
import tempfile
import httpx
from pydantic import BaseModel, Field
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


class Tools:
    class Valves(BaseModel):
        WEBUI_BASE_URL: str = Field(
            default="",
            description=(
                "Public-facing Open WebUI URL users access in their browser "
                "(e.g. http://192.168.1.100:8089). Leave blank to auto-detect from the request."
            ),
        )

    def __init__(self):
        self.valves = self.Valves()

    def _get_public_base_url(self, request) -> str:
        """URL for user-facing download links: valve > request.base_url > fallback."""
        if self.valves.WEBUI_BASE_URL:
            return self.valves.WEBUI_BASE_URL.rstrip("/")
        if request and hasattr(request, "base_url"):
            return str(request.base_url).rstrip("/")
        return "http://localhost:8080"

    def _get_internal_base_url(self) -> str:
        """URL for server-side API calls (always localhost inside the container)."""
        return "http://localhost:8080"

    def _get_token(self, request) -> str:
        if request and hasattr(request, "headers"):
            auth = request.headers.get("Authorization", "")
            if auth.startswith("Bearer "):
                return auth[7:]
        return ""

    async def create_spreadsheet(
        self,
        filename: str,
        sheets: str,
        include_totals: bool = True,
        __request__=None,
        __event_emitter__=None,
    ) -> str:
        """
        Create a formatted Excel spreadsheet (.xlsx) with one or more sheets.

        :param filename: Output filename without extension (e.g. 'budget_2026')
        :param sheets: Sheet data separated by '===SHEET:SheetName==='.
                       Within each sheet, rows are separated by '|' (preferred) or ',' (CSV).
                       Use '|' when values may contain commas (e.g. "1,000"). First row = header.
                       Example:
                       ===SHEET:Sales===
                       Month|Revenue|Expenses|Profit
                       January|50000|30000|20000
                       February|55000|32000|23000
                       ===SHEET:Summary===
                       Category|Value
                       Total Revenue|105000
        :param include_totals: Add a SUM row at the bottom of numeric columns
        :return: Download link and confirmation message
        """
        try:
            if __event_emitter__:
                await __event_emitter__(
                    {"type": "status", "data": {"description": "Building spreadsheet...", "done": False}}
                )

            wb = Workbook()
            wb.remove(wb.active)

            sheet_blocks = [b for b in sheets.split("===SHEET:") if b.strip()]

            for block in sheet_blocks:
                lines = block.strip().split("\n")
                sheet_name = lines[0].replace("===", "").strip()[:31]
                data_lines = [l for l in lines[1:] if l.strip()]
                if not data_lines:
                    continue

                ws = wb.create_sheet(title=sheet_name)
                delimiter = "|" if "|" in data_lines[0] else ","
                all_rows = [[c.strip() for c in l.split(delimiter)] for l in data_lines]
                headers = all_rows[0]
                data_rows = all_rows[1:]

                header_fill = PatternFill("solid", fgColor="1F4E79")
                header_font = Font(bold=True, color="FFFFFF", size=11)
                alt_fill = PatternFill("solid", fgColor="D6E4F0")
                total_fill = PatternFill("solid", fgColor="BDD7EE")
                total_font = Font(bold=True)
                thin = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )
                center = Alignment(horizontal="center", vertical="center")

                # Headers
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center
                    cell.border = thin
                ws.row_dimensions[1].height = 20

                # Data
                numeric_cols = set()
                for row_idx, row in enumerate(data_rows, 2):
                    fill = alt_fill if row_idx % 2 == 0 else PatternFill()
                    for col_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        try:
                            clean = (
                                value.replace(",", "").replace("$", "").replace("%", "")
                            )
                            num = float(clean)
                            cell.value = num
                            numeric_cols.add(col_idx)
                            if "$" in value:
                                cell.number_format = "$#,##0.00"
                            elif "%" in value:
                                cell.number_format = "0.00%"
                                cell.value = num / 100
                            else:
                                cell.number_format = "#,##0.##"
                        except (ValueError, AttributeError):
                            cell.value = value
                        cell.fill = fill
                        cell.border = thin
                        cell.alignment = Alignment(vertical="center")

                # Totals
                if include_totals and data_rows and numeric_cols:
                    total_row = len(data_rows) + 2
                    for col_idx in range(1, len(headers) + 1):
                        cell = ws.cell(row=total_row, column=col_idx)
                        if col_idx in numeric_cols:
                            col_letter = get_column_letter(col_idx)
                            cell.value = (
                                f"=SUM({col_letter}2:{col_letter}{total_row - 1})"
                            )
                            cell.number_format = "#,##0.##"
                        elif col_idx == 1:
                            cell.value = "TOTAL"
                        cell.fill = total_fill
                        cell.font = total_font
                        cell.border = thin

                # Auto-fit columns
                for col_idx, header in enumerate(headers, 1):
                    col_letter = get_column_letter(col_idx)
                    max_w = len(header) + 4
                    for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            if cell.value:
                                max_w = max(max_w, len(str(cell.value)) + 4)
                    ws.column_dimensions[col_letter].width = min(max_w, 50)

                ws.freeze_panes = "A2"

            # Save to temp and upload
            safe_name = re.sub(r"[^\w\-]", "_", filename).strip("_")
            xlsx_filename = f"{safe_name}.xlsx"

            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                tmp_path = tmp.name
                wb.save(tmp_path)

            if __event_emitter__:
                await __event_emitter__(
                    {"type": "status", "data": {"description": "Uploading...", "done": False}}
                )

            public_url = self._get_public_base_url(__request__)
            token = self._get_token(__request__)
            internal_url = self._get_internal_base_url()
            file_id = await self._upload_file(tmp_path, xlsx_filename, token, internal_url)
            os.unlink(tmp_path)

            sheet_names = [
                b.strip().split("\n")[0].replace("===", "").strip()
                for b in sheet_blocks
                if b.strip()
            ]
            download_url = f"{public_url}/api/v1/files/{file_id}/content/{xlsx_filename}"

            if __event_emitter__:
                await __event_emitter__(
                    {"type": "status", "data": {"description": "Done", "done": True}}
                )
                await __event_emitter__(
                    {
                        "type": "message",
                        "data": {"content": f"📥 **[Download {xlsx_filename}]({download_url})**"},
                    }
                )

            return (
                f"📥 [Download {xlsx_filename}]({download_url})\n\n"
                f"Spreadsheet created and uploaded (sheets: {', '.join(sheet_names)}). "
                f"Include this exact download link verbatim in your response."
            )

        except Exception as e:
            return f"❌ Spreadsheet creation failed: {str(e)}"

    async def _upload_file(
        self, file_path: str, filename: str, token: str, internal_url: str
    ) -> str:
        if not token:
            raise ValueError(
                "No Bearer token found in request. "
                "Check that __request__ is declared in the function signature."
            )
        headers = {"Authorization": f"Bearer {token}"}
        async with httpx.AsyncClient() as client:
            with open(file_path, "rb") as f:
                resp = await client.post(
                    f"{internal_url}/api/v1/files/",
                    headers=headers,
                    files={
                        "file": (
                            filename,
                            f,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
                    },
                    params={"process": "false"},
                    timeout=60,
                )
        resp.raise_for_status()
        return resp.json()["id"]
